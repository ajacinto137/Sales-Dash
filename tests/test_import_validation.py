"""Tests for the Excel user import's validation rules
(import_service.parse_and_validate()/commit_import()). No live appdb
needed -- user_store.find_sales_rep_by_name()/list_users() are
monkeypatched with in-memory fakes, since these tests are about the
validation/matching LOGIC, not the database layer underneath it (that's
exercised for real in the manual "Testing the Excel Import" walkthrough
in README.md).

Run with: pytest tests/test_import_validation.py -v
"""

import io

import openpyxl
import pytest

import import_service
import user_store


class _FakeFileStorage:
    """Minimal stand-in for Flask's request.files['file'] (a Werkzeug
    FileStorage) -- pandas.read_excel()/openpyxl need the full file-like
    protocol (read/seek/tell/...), so this proxies everything to a real
    io.BytesIO except `.filename`, which import_service.py reads to check
    the extension before ever touching pandas."""

    def __init__(self, buffer, filename):
        self._buffer = buffer
        self.filename = filename

    def __getattr__(self, name):
        return getattr(self._buffer, name)


def _make_xlsx(rows, filename="import.xlsx"):
    """rows: list of (rep_name, email, group) tuples. First row is
    always the header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Rep Name", "Email", "Group"])
    for row in rows:
        ws.append(list(row))
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return _FakeFileStorage(buffer, filename)


@pytest.fixture
def known_reps(monkeypatch):
    """Fakes sales_reps as {"Jack Jacinto", "Laryssa Dasilva"} and an
    empty existing-users table, both in-memory -- no appdb involved."""
    reps = {"jack jacinto": {"id": 1, "name": "Jack Jacinto"},
            "laryssa dasilva": {"id": 2, "name": "Laryssa Dasilva"}}

    def fake_find(name):
        return reps.get((name or "").strip().lower())

    monkeypatch.setattr(user_store, "find_sales_rep_by_name", fake_find)
    monkeypatch.setattr(user_store, "list_users", lambda: [])
    return reps


def test_valid_file_all_rows_importable(known_reps):
    file = _make_xlsx([
        ("Jack Jacinto", "jack@planet.net", "Sales Rep"),
        ("Laryssa Dasilva", "laryssa@planet.net", "Sales Rep"),
    ])
    ok, error, rows, summary = import_service.parse_and_validate(file)
    assert ok is True
    assert error is None
    assert summary["total"] == 2
    assert summary[import_service.ACTION_CANNOT_IMPORT] == 0
    assert all(r["proposed_action"] == import_service.ACTION_CREATE for r in rows)


def test_missing_required_column_rejected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Rep Name", "Email"])  # missing "Group"
    ws.append(["Jack Jacinto", "jack@planet.net"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file = _FakeFileStorage(buffer, "import.xlsx")

    ok, error, rows, summary = import_service.parse_and_validate(file)
    assert ok is False
    assert "Group" in error
    assert rows == []


def test_invalid_group_flagged_cannot_import(known_reps):
    file = _make_xlsx([("Jack Jacinto", "jack@planet.net", "Support")])
    ok, _error, rows, summary = import_service.parse_and_validate(file)
    assert ok is True
    assert rows[0]["proposed_action"] == import_service.ACTION_CANNOT_IMPORT
    assert 'Group "Support" is invalid' in rows[0]["errors"][0]


def test_blank_email_flagged_cannot_import(known_reps):
    file = _make_xlsx([("Jack Jacinto", "", "Sales Rep")])
    ok, _error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is True
    assert rows[0]["proposed_action"] == import_service.ACTION_CANNOT_IMPORT
    assert "Email is required" in rows[0]["errors"][0]


def test_malformed_email_flagged_cannot_import(known_reps):
    file = _make_xlsx([("Jack Jacinto", "not-an-email", "Sales Rep")])
    ok, _error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is True
    assert rows[0]["proposed_action"] == import_service.ACTION_CANNOT_IMPORT
    assert "Invalid email format" in rows[0]["errors"][0]


def test_duplicate_email_within_file_flagged(known_reps):
    file = _make_xlsx([
        ("Jack Jacinto", "same@planet.net", "Sales Rep"),
        ("Laryssa Dasilva", "same@planet.net", "Sales Rep"),
    ])
    ok, _error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is True
    assert all(r["proposed_action"] == import_service.ACTION_CANNOT_IMPORT for r in rows)
    assert all("Duplicate email" in r["errors"][0] for r in rows)


def test_unmatched_sales_rep_flagged_needs_review_not_cannot_import(known_reps):
    file = _make_xlsx([("Nobody Real", "nobody@planet.net", "Sales Rep")])
    ok, _error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is True
    assert rows[0]["proposed_action"] == import_service.ACTION_NEEDS_REVIEW
    assert rows[0]["matched_rep"] is None


def test_customer_success_row_not_flagged_needs_review_for_unmatched_rep(known_reps):
    """A Customer Success row's Rep Name is just a name, not a claim to a
    dashboard Sales Rep -- it should never need sales-rep matching."""
    file = _make_xlsx([("Some CS Person", "cs@planet.net", "Customer Success")])
    ok, _error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is True
    assert rows[0]["proposed_action"] == import_service.ACTION_CREATE


def test_empty_workbook_rejected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Rep Name", "Email", "Group"])  # header only, no data rows
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file = _FakeFileStorage(buffer, "import.xlsx")

    ok, error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is False
    assert "no data rows" in error.lower()


def test_unsupported_file_type_rejected():
    file = _FakeFileStorage(io.BytesIO(b"not a spreadsheet"), "import.csv")
    ok, error, rows, _summary = import_service.parse_and_validate(file)
    assert ok is False
    assert "Unsupported file type" in error


def test_partial_import_some_rows_succeed_some_fail(known_reps, monkeypatch):
    """Mirrors spec #10/#29: a bad row never blocks the good ones around
    it, and every failure carries its own specific reason."""
    created = []

    def fake_create_user(email, role, sales_rep_id=None):
        created.append((email, role, sales_rep_id))
        return True, None, {"id": len(created), "email": email}

    monkeypatch.setattr(user_store, "create_user", fake_create_user)
    monkeypatch.setattr(user_store, "update_user", lambda *a, **k: (True, None))

    file = _make_xlsx([
        ("Jack Jacinto", "jack@planet.net", "Sales Rep"),
        ("Bad Row", "bad@planet.net", "Support"),
        ("Laryssa Dasilva", "laryssa@planet.net", "Sales Rep"),
    ])
    _ok, _error, rows, _summary = import_service.parse_and_validate(file)

    result = import_service.commit_import(rows)
    assert result["added"] == 2
    assert len(result["failed"]) == 1
    assert result["failed"][0]["email"] == "bad@planet.net"
    assert "Support" in result["failed"][0]["reason"]
    assert len(created) == 2
